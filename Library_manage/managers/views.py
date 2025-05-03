from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Count, Sum, Q, Avg
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import datetime, timedelta
from account.models import User
from books.models import Book, BorrowingRecord, Category
from .forms import LibrarianForm, SystemSettingsForm, BorrowingSettingsForm, FineSettingsForm
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from reviews.models import Review
from .models import Fines, Analytics, Budget, SystemSettings, UserRole
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.conf import settings
import plotly.express as px
import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font

def is_manager(user):
    """Check if user is a manager"""
    return user.is_authenticated and user.role == 'MANAGER'

def is_admin(user):
    if not user.is_authenticated:
        return False
    return user.role == 'MANAGER'  # In this system, managers are also admins

@login_required
@user_passes_test(is_manager)
def home(request):
    """Manager home page with quick overview and navigation"""
    # Get basic statistics
    total_books = Book.objects.count()
    total_borrowed = BorrowingRecord.objects.filter(status='BORROWED').count()
    overdue_books = BorrowingRecord.objects.filter(
        status='BORROWED',
        due_date__lt=timezone.now()
    ).count()
    total_librarians = User.objects.filter(userrole__role='librarian').count()
    total_users = User.objects.count()

    # Get recent activity
    recent_activity = BorrowingRecord.objects.select_related('user', 'book').order_by('-borrow_date')[:5]
    
    context = {
        'total_books': total_books,
        'total_borrowed': total_borrowed,
        'overdue_books': overdue_books,
        'total_librarians': total_librarians,
        'total_users': total_users,
        'recent_activity': recent_activity,
    }
    return render(request, 'managers/home.html', context)

@login_required
@user_passes_test(is_manager)
def dashboard(request):
    """Manager dashboard with management features"""
    # Get recent activity
    recent_activity = BorrowingRecord.objects.select_related('book', 'borrower').order_by('-borrowed_date')[:10]
    
    # Get system settings
    system_settings = SystemSettings.objects.first()
    if not system_settings:
        system_settings = SystemSettings.objects.create()
    
    # Get basic statistics
    total_books = Book.objects.count()
    total_borrowed = BorrowingRecord.objects.filter(status='BORROWED').count()
    overdue_books = BorrowingRecord.objects.filter(
        status='BORROWED',
        due_date__lt=timezone.now()
    ).count()
    total_librarians = User.objects.filter(role='LIBRARIAN').count()
    total_users = User.objects.count()
    
    context = {
        'recent_activity': recent_activity,
        'system_settings': system_settings,
        'total_books': total_books,
        'total_borrowed': total_borrowed,
        'overdue_books': overdue_books,
        'total_librarians': total_librarians,
        'total_users': total_users,
    }
    return render(request, 'managers/dashboard.html', context)

@login_required
@user_passes_test(is_admin)
def admin_control_panel(request):
    system_settings = SystemSettings.objects.first()
    if not system_settings:
        system_settings = SystemSettings.objects.create()
    
    user_roles = UserRole.objects.all()
    
    context = {
        'system_settings': system_settings,
        'user_roles': user_roles,
    }
    return render(request, 'managers/admin_panel.html', context)

@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def toggle_feature(request):
    feature = request.POST.get('feature')
    enabled = request.POST.get('enabled') == 'true'
    
    settings = SystemSettings.objects.first()
    if not settings:
        settings = SystemSettings.objects.create()
    
    if feature == 'review_system':
        settings.review_system_enabled = enabled
    elif feature == 'google_books_api':
        settings.google_books_api_enabled = enabled
    
    settings.save()
    return JsonResponse({'status': 'success'})

@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def update_system_settings(request):
    try:
        data = json.loads(request.body)
        settings = SystemSettings.objects.first()
        if not settings:
            settings = SystemSettings.objects.create()

        # Update settings based on the received data
        if 'review_system' in data:
            settings.review_system_enabled = data['review_system']
        if 'google_books_api' in data:
            settings.google_books_api_enabled = data['google_books_api']
        if 'maintenance_message' in data:
            settings.maintenance_message = data['maintenance_message']
        if 'maintenance_date' in data:
            try:
                maintenance_date = datetime.strptime(data['maintenance_date'], '%Y-%m-%dT%H:%M')
                settings.maintenance_scheduled = maintenance_date
            except ValueError:
                settings.maintenance_scheduled = None

        settings.save()
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@login_required
@user_passes_test(is_manager)
def export_report(request, report_type):
    if report_type == 'borrowings':
        data = BorrowingRecord.objects.all().values(
            'book__title',
            'borrower__username',
            'borrowed_date',
            'return_date',
            'due_date'
        )
        df = pd.DataFrame(data)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="borrowing_report.csv"'
        df.to_csv(response, index=False)
        return response
    
    elif report_type == 'inventory':
        data = Book.objects.all().values(
            'title',
            'author',
            'isbn',
            'total_copies',
            'available_copies'
        )
        df = pd.DataFrame(data)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="inventory_report.csv"'
        df.to_csv(response, index=False)
        return response
    
    elif report_type == 'users':
        data = User.objects.all().values(
            'username',
            'email',
            'first_name',
            'last_name',
            'role',
            'is_active',
            'date_joined',
            'last_login'
        )
        df = pd.DataFrame(data)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="users_report.csv"'
        df.to_csv(response, index=False)
        return response
    
    elif report_type == 'fines':
        data = BorrowingRecord.objects.filter(fine_amount__gt=0).values(
            'book__title',
            'borrower__username',
            'borrowed_date',
            'due_date',
            'return_date',
            'fine_amount',
            'status'
        )
        df = pd.DataFrame(data)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="fines_report.csv"'
        df.to_csv(response, index=False)
        return response
    
    return HttpResponse('Invalid report type', status=400)

@login_required
@user_passes_test(is_admin)
def manage_user_roles(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        role = request.POST.get('role')
        permissions = json.loads(request.POST.get('permissions', '{}'))
        
        user = User.objects.get(id=user_id)
        user_role, created = UserRole.objects.get_or_create(user=user)
        user_role.role = role
        user_role.permissions = permissions
        user_role.save()
        
        messages.success(request, 'User role updated successfully')
        return redirect('admin_control_panel')
    
    users = User.objects.all()
    context = {
        'users': users,
    }
    return render(request, 'managers/manage_roles.html', context)

@login_required
@user_passes_test(is_manager)
def manage_librarians(request):
    librarians = User.objects.filter(role='LIBRARIAN')
    return render(request, 'managers/manage_librarians.html', {'librarians': librarians})

@login_required
@user_passes_test(is_manager)
def add_librarian(request):
    if request.method == 'POST':
        form = LibrarianForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_librarian = True
            user.save()
            messages.success(request, 'Librarian added successfully!')
            return redirect('managers:manage_librarians')
    else:
        form = LibrarianForm()
    return render(request, 'managers/librarian_form.html', {'form': form, 'title': 'Add Librarian'})

@login_required
@user_passes_test(is_manager)
def edit_librarian(request, pk):
    librarian = get_object_or_404(User, pk=pk, role='LIBRARIAN')
    if request.method == 'POST':
        form = LibrarianForm(request.POST, instance=librarian)
        if form.is_valid():
            form.save()
            messages.success(request, 'Librarian updated successfully!')
            return redirect('managers:manage_librarians')
    else:
        form = LibrarianForm(instance=librarian)
    return render(request, 'managers/librarian_form.html', {'form': form, 'title': 'Edit Librarian'})

@login_required
@user_passes_test(is_manager)
def delete_librarian(request, pk):
    librarian = get_object_or_404(User, pk=pk, role='LIBRARIAN')
    if request.method == 'POST':
        librarian.delete()
        messages.success(request, 'Librarian deleted successfully!')
        return redirect('managers:manage_librarians')
    return render(request, 'managers/confirm_delete.html', {'object': librarian})

@login_required
@user_passes_test(is_manager)
def system_settings(request):
    """System settings management view"""
    settings = SystemSettings.objects.first()
    if not settings:
        settings = SystemSettings.objects.create()
    
    if request.method == 'POST':
        form = SystemSettingsForm(request.POST, instance=settings)
        if form.is_valid():
            form.save()
            messages.success(request, 'System settings updated successfully!')
            return redirect('managers:dashboard')
    else:
        form = SystemSettingsForm(instance=settings)
    
    context = {
        'form': form,
        'settings': settings,
    }
    return render(request, 'managers/system_settings.html', context)

@login_required
@user_passes_test(is_manager)
def borrowing_settings(request):
    if request.method == 'POST':
        form = BorrowingSettingsForm(request.POST)
        if form.is_valid():
            # Save borrowing settings logic here
            messages.success(request, 'Borrowing settings updated successfully!')
            return redirect('managers:borrowing_settings')
    else:
        form = BorrowingSettingsForm()
    return render(request, 'managers/borrowing_settings.html', {'form': form})

@login_required
@user_passes_test(is_manager)
def fine_settings(request):
    if request.method == 'POST':
        form = FineSettingsForm(request.POST)
        if form.is_valid():
            # Save fine settings logic here
            messages.success(request, 'Fine settings updated successfully!')
            return redirect('managers:fine_settings')
    else:
        form = FineSettingsForm()
    return render(request, 'managers/fine_settings.html', {'form': form})

@login_required
@user_passes_test(is_manager)
def borrowing_report(request):
    # Get borrowing statistics
    total_borrowings = BorrowingRecord.objects.count()
    active_borrowings = BorrowingRecord.objects.filter(status='BORROWED').count()
    overdue_borrowings = BorrowingRecord.objects.filter(
        status='BORROWED',
        due_date__lt=timezone.now()
    ).count()
    
    # Get popular books
    popular_books = Book.objects.annotate(
        borrow_count=Count('borrowing_records')
    ).order_by('-borrow_count')[:5]
    
    context = {
        'total_borrowings': total_borrowings,
        'active_borrowings': active_borrowings,
        'overdue_borrowings': overdue_borrowings,
        'popular_books': popular_books,
    }
    return render(request, 'managers/borrowing_report.html', context)

@login_required
@user_passes_test(is_manager)
def fine_report(request):
    # Get fine statistics
    total_fines = BorrowingRecord.objects.aggregate(
        total=Sum('fine_amount')
    )['total'] or 0
    
    unpaid_fines = BorrowingRecord.objects.filter(
        fine_amount__gt=0,
        status__in=['BORROWED', 'OVERDUE']
    ).aggregate(total=Sum('fine_amount'))['total'] or 0
    
    context = {
        'total_fines': total_fines,
        'unpaid_fines': unpaid_fines,
    }
    return render(request, 'managers/fine_report.html', context)

@login_required
@user_passes_test(is_manager)
def inventory_report(request):
    # Get inventory statistics
    total_books = Book.objects.count()
    available_books = Book.objects.filter(available_copies__gt=0).count()
    categories = Book.objects.values('category__name').annotate(
        count=Count('id')
    )
    
    context = {
        'total_books': total_books,
        'available_books': available_books,
        'categories': categories,
    }
    return render(request, 'managers/inventory_report.html', context)

@login_required
@user_passes_test(is_manager)
def manage_users(request):
    users = User.objects.all().prefetch_related('book_borrowings')
    for user in users:
        user.borrowed_count = user.book_borrowings.filter(status='BORROWED').count()
    return render(request, 'managers/manage_users.html', {'users': users})

@login_required
@user_passes_test(is_manager)
def user_detail(request, pk):
    user = get_object_or_404(User, pk=pk)
    borrowing_history = BorrowingRecord.objects.filter(borrower=user)
    return render(request, 'managers/user_detail.html', {
        'user': user,
        'borrowing_history': borrowing_history
    })

@login_required
@user_passes_test(is_manager)
def suspend_user(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user.is_active = not user.is_active
        user.save()
        action = 'suspended' if not user.is_active else 'activated'
        messages.success(request, f'User {action} successfully!')
        return redirect('managers:user_detail', pk=pk)
    return render(request, 'managers/confirm_suspend.html', {'user': user})

@login_required
@user_passes_test(is_manager)
def confirm_suspend(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user.is_active = not user.is_active
        user.save()
        action = 'suspended' if not user.is_active else 'activated'
        messages.success(request, f'User {action} successfully!')
        return redirect('managers:manage_users')
    
    context = {
        'user': user,
        'action': 'suspend' if user.is_active else 'activate'
    }
    return render(request, 'managers/confirm_suspend.html', context)

class ReportsView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    template_name = 'managers/reports.html'
    context_object_name = 'reports'

    def test_func(self):
        return self.request.user.is_manager

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Book statistics
        context['total_books'] = Book.objects.count()
        context['books_by_category'] = Category.objects.annotate(book_count=Count('books'))
        context['top_rated_books'] = Book.objects.annotate(
            avg_rating=Avg('reviews__rating')
        ).order_by('-avg_rating')[:5]
        
        # User statistics
        context['total_users'] = User.objects.count()
        context['active_users'] = User.objects.filter(is_active=True).count()
        
        # Review statistics
        context['total_reviews'] = Review.objects.count()
        context['reviews_by_month'] = Review.objects.filter(
            created_at__gte=timezone.now() - timedelta(days=30)
        ).count()
        
        # Fine statistics using BorrowingRecord
        context['total_fines'] = BorrowingRecord.objects.aggregate(
            total=Sum('fine_amount')
        )['total'] or 0
        context['unpaid_fines'] = BorrowingRecord.objects.filter(
            fine_amount__gt=0,
            status__in=['BORROWED', 'OVERDUE']
        ).aggregate(total=Sum('fine_amount'))['total'] or 0
        
        return context

    def get_queryset(self):
        # Return an empty queryset since we're using context data
        return []

@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def update_user_role(request):
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        role = data.get('role')
        permissions = data.get('permissions', {})

        if not user_id or not role:
            return JsonResponse({'status': 'error', 'message': 'Missing required fields'}, status=400)

        user = get_object_or_404(User, id=user_id)
        user_role, created = UserRole.objects.get_or_create(user=user)
        
        user_role.role = role
        user_role.permissions = permissions
        user_role.save()

        return JsonResponse({'status': 'success'})
    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@login_required
@user_passes_test(is_manager)
def export_borrowing_report(request):
    # Get all borrowing records with related book and borrower information
    borrowing_records = BorrowingRecord.objects.select_related('book', 'borrower').all()
    
    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Borrowing Report"
    
    # Add headers
    headers = ['Book Title', 'Borrower', 'Borrowed Date', 'Due Date', 'Return Date', 'Status']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Add data
    for row_num, record in enumerate(borrowing_records, 2):
        ws.cell(row=row_num, column=1, value=record.book.title)
        ws.cell(row=row_num, column=2, value=f"{record.borrower.first_name} {record.borrower.last_name}")
        ws.cell(row=row_num, column=3, value=record.borrowed_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=4, value=record.due_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=5, value=record.return_date.strftime('%Y-%m-%d') if record.return_date else '')
        ws.cell(row=row_num, column=6, value=record.get_status_display())
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=borrowing_report.xlsx'
    wb.save(response)
    
    return response
